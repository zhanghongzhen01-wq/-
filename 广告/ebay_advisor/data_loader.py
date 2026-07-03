import pandas as pd
import numpy as np


def load_sales_data(file_path):
    df = pd.read_excel(file_path)
    return df


def load_ad_data(file_path):
    """加载CPC广告数据，正确处理多段式Excel结构"""
    import openpyxl

    # 关键词段落的标准列名
    KW_HEADERS = [
        'Ad group name', 'Keyword', 'Match type', 'My bid', 'Status',
        'Impressions', 'Clicks', 'Click through rate', 'Sales',
        'Conversion rate', 'Sale amount', 'Ad Fees', 'ROAS', 'ACOS',
        'ROI', 'Avg cost per sale', 'Cost per click'
    ]

    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        campaign_name = sheet_name

        # 扫描找关键词段落（扩大范围到 40 行，覆盖所有可能的 Sheet 结构）
        kw_start = None
        for r in range(1, min(ws.max_row + 1, 40)):
            cell_val = str(ws.cell(row=r, column=1).value or '').strip()
            if cell_val == 'Ad group name':
                col3 = str(ws.cell(row=r, column=3).value or '').strip()
                # 只有第3列是 Match type 才是关键词段落，第2列是 Item# 的是listing段落
                if col3 == 'Match type':
                    kw_start = r
                    break

        if kw_start is None:
            continue

        # 读取关键词行
        for r in range(kw_start + 1, ws.max_row + 1):
            kw = str(ws.cell(row=r, column=2).value or '').strip()
            if not kw:
                continue
            # 过滤eBay系统汇总行
            if kw.lower().startswith('consolidated data from'):
                continue

            row_data = {'Campaign': campaign_name}
            for ci, h in enumerate(KW_HEADERS):
                val = ws.cell(row=r, column=ci + 1).value
                if val is None:
                    row_data[h] = np.nan
                elif isinstance(val, str):
                    val = val.strip()
                    if val.endswith(' USD'):
                        try:
                            val = float(val.replace(' USD', '').replace(',', ''))
                        except ValueError:
                            pass
                    elif val.endswith('%'):
                        try:
                            val = float(val.replace('%', '')) / 100
                        except ValueError:
                            pass
                    else:
                        # 尝试转换纯数字字符串
                        try:
                            if '.' in val:
                                val = float(val)
                            else:
                                val = int(val)
                        except ValueError:
                            pass
                    row_data[h] = val
                else:
                    row_data[h] = val

            all_rows.append(row_data)

    wb.close()
    df = pd.DataFrame(all_rows)

    # 补充缺失的列
    for h in KW_HEADERS:
        if h not in df.columns:
            df[h] = np.nan

    # 数值列补充默认值并清理脏数据
    for col in ['Clicks', 'Sales', 'Impressions', 'Ad Fees', 'Sale amount', 'ROAS', 'Click through rate', 'Conversion rate']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: np.nan if isinstance(x, str) and x.strip() == '' else x)
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    # ACOS 单独处理：保留 NaN，不填0（0 会误导为高效）
    if 'ACOS' in df.columns:
        df['ACOS'] = df['ACOS'].apply(lambda x: np.nan if isinstance(x, str) and x.strip() == '' else x)
        df['ACOS'] = pd.to_numeric(df['ACOS'], errors='coerce')

    print(f"  关键词行数: {len(df)}")
    return df


def load_item_sku_keywords(file_path):
    """解析CPC数据，建立 SKU ↔ 关键词 映射。
    返回 DataFrame: Campaign, Ad group name, Keyword, Match type, My bid, Status,
                   Clicks, Sales, ACOS, ROAS, Ad Fees, Impressions,
                   Item#, SKU
    """
    import openpyxl

    KW_HEADERS = [
        'Ad group name', 'Keyword', 'Match type', 'My bid', 'Status',
        'Impressions', 'Clicks', 'Click through rate', 'Sales',
        'Conversion rate', 'Sale amount', 'Ad Fees', 'ROAS', 'ACOS',
        'ROI', 'Avg cost per sale', 'Cost per click'
    ]

    ITEM_HEADERS = [
        'Ad group name', 'Item#', 'Item title', 'SKU', 'ListingType',
        'QTY', 'Sold', 'eBay counter', 'eBay watch',
        'eBay suggested listing', 'Status', 'Impressions', 'Clicks',
        'Click through rate', 'Sales', 'Conversion rate', 'Sale amount',
        'Ad Fees', 'ROAS', 'ACOS', 'ROI', 'Avg cost per sale', 'Cost per click'
    ]

    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        campaign_name = sheet_name

        # 找关键词段落
        kw_start = None
        item_start = None
        for r in range(1, min(ws.max_row + 1, 40)):
            cell_val = str(ws.cell(row=r, column=1).value or '').strip()
            if cell_val == 'Ad group name':
                col2 = str(ws.cell(row=r, column=2).value or '').strip()
                col3 = str(ws.cell(row=r, column=3).value or '').strip()
                if col3 == 'Match type':
                    kw_start = r
                elif col2 == 'Item#' and item_start is None:
                    item_start = r

        if kw_start is None:
            continue

        # 建立 Ad group name → [(Item#, SKU)] 映射
        ad_group_items = {}
        if item_start:
            for r in range(item_start + 1, ws.max_row + 1):
                ag = str(ws.cell(row=r, column=1).value or '').strip()
                item_num = str(ws.cell(row=r, column=2).value or '').strip()
                sku = str(ws.cell(row=r, column=4).value or '').strip()
                if not ag or not item_num:
                    break  # End of item section
                if ag not in ad_group_items:
                    ad_group_items[ag] = []
                ad_group_items[ag].append((item_num, sku))

        # 读关键词，关联 Ad group → Items
        for r in range(kw_start + 1, ws.max_row + 1):
            kw = str(ws.cell(row=r, column=2).value or '').strip()
            if not kw:
                continue
            if kw.lower().startswith('consolidated data from'):
                continue

            ag = str(ws.cell(row=r, column=1).value or '').strip()
            items_in_group = ad_group_items.get(ag, [('', '')])

            for item_num, sku in items_in_group:
                row_data = {'Campaign': campaign_name, 'Item#': item_num, 'SKU': sku}
                for ci, h in enumerate(KW_HEADERS):
                    val = ws.cell(row=r, column=ci + 1).value
                    if val is None:
                        row_data[h] = np.nan
                    elif isinstance(val, str):
                        val = val.strip()
                        if val.endswith(' USD'):
                            try: val = float(val.replace(' USD', '').replace(',', ''))
                            except ValueError: pass
                        elif val.endswith('%'):
                            try: val = float(val.replace('%', '')) / 100
                            except ValueError: pass
                        else:
                            try: val = float(val) if '.' in val else int(val)
                            except ValueError: pass
                        row_data[h] = val
                    else:
                        row_data[h] = val
                all_rows.append(row_data)

    wb.close()
    df = pd.DataFrame(all_rows)
    for col in ['Clicks','Sales','Impressions','Ad Fees','Sale amount','ROAS']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: np.nan if isinstance(x, str) and x.strip() == '' else x)
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    if 'ACOS' in df.columns:
        df['ACOS'] = df['ACOS'].apply(lambda x: np.nan if isinstance(x, str) and x.strip() == '' else x)
        df['ACOS'] = pd.to_numeric(df['ACOS'], errors='coerce')

    print(f"  SKU-关键词映射行数: {len(df)}")
    return df


def _clean_ad_data(df):
    """保留作为兼容，load_ad_data 已内联处理"""
    return df
