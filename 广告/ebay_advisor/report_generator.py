import pandas as pd
from datetime import datetime


class ReportGenerator:
    def __init__(self):
        self.report = ""
    
    def generate_full_report(self, sku_evaluator, keyword_optimizer, potential_predictor):
        self.report = ""
        
        self._add_header()
        self._add_methodology()
        self._add_sku_recommendations(sku_evaluator)
        self._add_keyword_optimization(keyword_optimizer)
        self._add_potential_prediction(potential_predictor)
        self._add_summary()
        
        return self.report
    
    def _add_header(self):
        self.report += "=" * 70 + "\n"
        self.report += "          eBay广告投放智能分析报告          \n"
        self.report += "=" * 70 + "\n"
        self.report += f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        self.report += "-" * 70 + "\n\n"
    
    def _add_methodology(self):
        self.report += "【一、eBay广告投放方法论】\n"
        self.report += "-" * 70 + "\n\n"
        
        self.report += "1.1 广告类型选择\n"
        self.report += "=" * 30 + "\n"
        self.report += "  PLG / Promoted Listings General (Cost Per Sale):\n"
        self.report += "    - 按成交收费，仅在产生订单时付费（原eBay Standard）\n"
        self.report += "    - 优点：风险低，不出单不花钱\n"
        self.report += "    - 缺点：eBay佣金费率较高(通常12-15%)\n"
        self.report += "    - 适用场景：成熟商品(转化率>2%)、爆款商品\n"
        self.report += "\n"
        self.report += "  PLP / Promoted Listings Advanced (Cost Per Click):\n"
        self.report += "    - 按点击收费，适合新品测试、关键词探索（原eBay Advanced）\n"
        self.report += "    - 优点：成本可控，能快速获取流量、抢占头部位置\n"
        self.report += "    - 缺点：可能产生无效点击，需要持续优化关键词\n"
        self.report += "    - 适用场景：新品期(0-30天)、需要快速曝光、精准关键词投放\n"
        self.report += "\n"
        
        self.report += "1.2 预算分配策略\n"
        self.report += "=" * 30 + "\n"
        self.report += "  A级商品(高优先级)：30%预算\n"
        self.report += "    - 特征：销量稳定、增长迅猛、ROAS优良\n"
        self.report += "    - 策略：重点投放，最大化产出\n"
        self.report += "\n"
        self.report += "  B级商品(中优先级)：40%预算\n"
        self.report += "    - 特征：销量中等、增长稳定\n"
        self.report += "    - 策略：持续投放，稳步提升\n"
        self.report += "\n"
        self.report += "  C级商品(低优先级)：20%预算\n"
        self.report += "    - 特征：销量较低、增长缓慢\n"
        self.report += "    - 策略：维持基础投放，观察变化\n"
        self.report += "\n"
        self.report += "  D/E级商品：10%预算\n"
        self.report += "    - 特征：无销量或负增长\n"
        self.report += "    - 策略：少量测试或暂停\n"
        self.report += "\n"
        
        self.report += "1.3 关键词优化方法\n"
        self.report += "=" * 30 + "\n"
        self.report += "  关键词匹配方式：\n"
        self.report += "    - 广泛匹配：覆盖更多搜索词，适合新品拓词\n"
        self.report += "    - 词组匹配：精准度适中，平衡流量与转化\n"
        self.report += "    - 精确匹配：最精准，适合高转化词\n"
        self.report += "\n"
        self.report += "  关键词筛选标准：\n"
        self.report += "    - 点击>50且无转化：直接删除\n"
        self.report += "    - ACoS<30%：扩大投放\n"
        self.report += "    - ACoS 30-50%：保持观察\n"
        self.report += "    - ACoS>50%：降价或暂停\n"
        self.report += "\n"
        
        self.report += "1.4 数据监控指标\n"
        self.report += "=" * 30 + "\n"
        self.report += "  CTR (点击率)：点击数/曝光数，反映关键词质量\n"
        self.report += "  Conversion Rate (转化率)：订单数/点击数，反映落地页质量\n"
        self.report += "  ACoS (广告销售成本比)：广告费/广告销售额，反映广告效率\n"
        self.report += "  ROAS (广告支出回报率)：广告销售额/广告费，ACoS的倒数\n"
        self.report += "  TACoS (总广告销售成本比)：广告费/总销售额，反映整体盈利状况\n"
        self.report += "\n"
        
        self.report += "1.5 进阶优化技巧\n"
        self.report += "=" * 30 + "\n"
        self.report += "  - 分时投放：根据销售高峰调整出价\n"
        self.report += "  - 地域优化：重点投放高转化地区\n"
        self.report += "  - 否定关键词：排除无效搜索词\n"
        self.report += "  - 动态出价：利用eBay自动出价功能\n"
        self.report += "  - 商品组合：搭配销售，提升客单价\n"
        self.report += "\n\n"
    
    def _add_sku_recommendations(self, evaluator):
        self.report += "【二、SKU投放价值评估】\n"
        self.report += "-" * 70 + "\n\n"
        
        top_skus = evaluator.get_top_skus(15)
        self.report += f"2.1 高优先级SKU推荐 (共{len(top_skus)}个)\n"
        self.report += "=" * 30 + "\n"
        
        self.report += f"{'排名':<5} {'SKU':<12} {'15天销量':<10} {'15天销售额':<12} {'投放评分':<8} {'优先级':<6} {'推荐广告':<12} {'建议预算':<8}\n"
        self.report += "-" * 80 + "\n"
        
        for i, (_, row) in enumerate(top_skus.iterrows(), 1):
            self.report += f"{i:<5} {row['SKU']:<12} {row['15天销量']:<10} {row['15天销售额']:<12.2f} {row['投放评分']:<8} {row['投放优先级']:<6} {row['推荐广告类型']:<12} {row['建议预算占比']*100:<6.1f}%\n"
        
        self.report += "\n"
        
        low_skus = evaluator.get_low_priority_skus()
        self.report += f"2.2 低优先级SKU (共{len(low_skus)}个)\n"
        self.report += "=" * 30 + "\n"
        
        if len(low_skus) > 0:
            self.report += f"{'SKU':<12} {'15天销量':<10} {'销量环比':<10} {'建议':<10}\n"
            self.report += "-" * 45 + "\n"
            
            for _, row in low_skus.head(10).iterrows():
                self.report += f"{row['SKU']:<12} {row['15天销量']:<10} {row['销量环比']:<10} 暂停投放\n"
            
            if len(low_skus) > 10:
                self.report += f"... 还有{len(low_skus)-10}个低优先级SKU\n"
        else:
            self.report += "暂无低优先级SKU\n"
        
        self.report += "\n\n"
    
    def _add_keyword_optimization(self, optimizer):
        self.report += "【三、广告词组优化建议】\n"
        self.report += "-" * 70 + "\n\n"
        
        shutdown_list = optimizer.get_shutdown_list(limit=300)
        self.report += f"3.1 建议空烧关停的广告词组 (共{len(shutdown_list)}个，展示前300)\n"
        self.report += "=" * 30 + "\n"
        
        if len(shutdown_list) > 0:
            self.report += f"{'序号':<5} {'广告组':<18} {'关键词':<22} {'类型':<7} {'点击':<7} {'ACoS':<7} {'原因':<30}\n"
            self.report += "-" * 95 + "\n"
            
            for i, (_, row) in enumerate(shutdown_list.iterrows(), 1):
                acos_val = f"{row['ACOS']:.1%}" if pd.notna(row['ACOS']) else "-"
                ad_group = str(row.get('Campaign', ''))[:18]
                match_type = str(row.get('Match type', ''))[:7]
                self.report += f"{i:<5} {ad_group:<18} {str(row['Keyword'])[:22]:<22} {match_type:<7} {row['Clicks']:<7} {acos_val:<7} {row['优化建议'][:30]:<30}\n"
        else:
            self.report += "暂无需要关停的广告词组\n"
        
        self.report += "\n"
        
        expand_list = optimizer.get_expand_list()
        self.report += f"3.2 建议追加预算的广告词组 (共{len(expand_list)}个)\n"
        self.report += "=" * 30 + "\n"
        
        if len(expand_list) > 0:
            self.report += f"{'序号':<5} {'广告组':<18} {'关键词':<22} {'类型':<7} {'点击':<7} {'转化':<7} {'ACoS':<7} {'ROAS':<7} {'建议':<30}\n"
            self.report += "-" * 100 + "\n"
            
            for i, (_, row) in enumerate(expand_list.iterrows(), 1):
                acos_val = f"{row['ACOS']:.1%}" if pd.notna(row['ACOS']) else "-"
                roas_val = f"{row['ROAS']:.1f}" if pd.notna(row['ROAS']) else "-"
                ad_group = str(row.get('Campaign', ''))[:18]
                match_type = str(row.get('Match type', ''))[:7]
                self.report += f"{i:<5} {ad_group:<18} {str(row['Keyword'])[:22]:<22} {match_type:<7} {row['Clicks']:<7} {row['Sales']:<7} {acos_val:<7} {roas_val:<7} {row['优化建议'][:30]:<30}\n"
        else:
            self.report += "暂无需要扩量的广告词组\n"
        
        self.report += "\n"
        
        optimize_list = optimizer.get_optimize_list()
        self.report += f"3.3 建议降价的广告词组 (共{len(optimize_list)}个)\n"
        self.report += "=" * 30 + "\n"
        
        if len(optimize_list) > 0:
            self.report += f"{'序号':<5} {'广告组':<18} {'关键词':<22} {'类型':<7} {'点击':<7} {'转化':<7} {'ACoS':<8} {'建议':<30}\n"
            self.report += "-" * 100 + "\n"
            
            for i, (_, row) in enumerate(optimize_list.iterrows(), 1):
                acos_val = f"{row['ACOS']:.1%}" if pd.notna(row['ACOS']) else "-"
                ad_group = str(row.get('Campaign', ''))[:18]
                match_type = str(row.get('Match type', ''))[:7]
                self.report += f"{i:<5} {ad_group:<18} {str(row['Keyword'])[:22]:<22} {match_type:<7} {row['Clicks']:<7} {row['Sales']:<7} {acos_val:<8} {row['优化建议'][:30]:<30}\n"
        else:
            self.report += "暂无需要降价的广告词组\n"
        
        self.report += "\n"
        
        suggested_keywords = optimizer.get_suggested_keywords()
        self.report += f"3.4 建议新增的广告词组 (共{len(suggested_keywords)}个)\n"
        self.report += "=" * 30 + "\n"
        
        if len(suggested_keywords) > 0:
            self.report += f"{'序号':<5} {'原始关键词':<20} {'建议关键词':<30} {'原因':<30}\n"
            self.report += "-" * 90 + "\n"
            
            for i, (_, row) in enumerate(suggested_keywords.head(15).iterrows(), 1):
                self.report += f"{i:<5} {str(row['原始关键词'])[:20]:<20} {str(row['建议关键词'])[:30]:<30} {row['建议原因']:<30}\n"
            
            if len(suggested_keywords) > 15:
                self.report += f"... 还有{len(suggested_keywords)-15}个建议关键词\n"
        else:
            self.report += "暂无建议新增的关键词\n"
        
        self.report += "\n\n"
    
    def _add_potential_prediction(self, predictor):
        self.report += "【四、SKU投放潜力预判】\n"
        self.report += "-" * 70 + "\n\n"
        
        high_potential = predictor.get_high_potential_skus()
        self.report += f"4.1 有潜力的SKU (共{len(high_potential)}个)\n"
        self.report += "=" * 30 + "\n"
        
        if len(high_potential) > 0:
            self.report += f"{'序号':<5} {'SKU':<12} {'15天销量':<10} {'销量环比':<10} {'潜力评分':<8} {'潜力等级':<8} {'分析':<40}\n"
            self.report += "-" * 90 + "\n"
            
            for i, (_, row) in enumerate(high_potential.head(15).iterrows(), 1):
                self.report += f"{i:<5} {row['SKU']:<12} {row['15天销量']:<10} {row['销量环比']:<10} {row['潜力评分']:<8} {row['潜力等级']:<8} {row['潜力分析'][:40]:<40}\n"
            
            if len(high_potential) > 15:
                self.report += f"... 还有{len(high_potential)-15}个有潜力SKU\n"
        else:
            self.report += "暂无高潜力SKU\n"
        
        self.report += "\n\n"
    
    def _add_summary(self):
        self.report += "【五、总结与建议】\n"
        self.report += "-" * 70 + "\n\n"
        
        self.report += "5.1 行动清单\n"
        self.report += "=" * 20 + "\n"
        self.report += "  ▶ 立即执行：\n"
        self.report += "    - 关停无点击/无转化的广告词组\n"
        self.report += "    - 对高优先级SKU增加广告预算\n"
        self.report += "\n"
        self.report += "  ▶ 本周执行：\n"
        self.report += "    - 测试新推荐的广告词组\n"
        self.report += "    - 关注新品机会的投放效果\n"
        self.report += "\n"
        self.report += "  ▶ 持续优化：\n"
        self.report += "    - 每日查看广告数据变化\n"
        self.report += "    - 根据ACoS调整出价策略\n"
        self.report += "\n"
        
        self.report += "5.2 新手入门建议\n"
        self.report += "=" * 20 + "\n"
        self.report += "  1. 从PLP开始：先用按点击付费测试关键词和商品\n"
        self.report += "  2. 小预算起步：每个SKU每天预算控制在$5-$10\n"
        self.report += "  3. 密切监控：每天查看广告数据，及时调整\n"
        self.report += "  4. 关键词精选：从广泛匹配开始，逐步筛选优质关键词\n"
        self.report += "  5. 耐心测试：广告效果需要时间积累，不要急于否定\n"
        self.report += "\n"
        
        self.report += "5.3 常见误区\n"
        self.report += "=" * 20 + "\n"
        self.report += "  × 不要给所有商品都投广告，重点培养爆款\n"
        self.report += "  × 不要均匀分配预算，集中资源在高潜力商品\n"
        self.report += "  × 不要忽略ACoS，超过50%的广告需要优化\n"
        self.report += "  × 不要频繁调整出价，给数据留足积累时间\n"
        self.report += "  × 不要只看短期数据，建议观察至少7天\n"
        self.report += "\n"
        
        self.report += "=" * 70 + "\n"
        self.report += "              报告结束 - 祝您广告投放顺利！              \n"
        self.report += "=" * 70 + "\n"
    
    def save_report(self, file_path):
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(self.report)
        return file_path

    def save_excel(self, file_path, sku_evaluator, keyword_optimizer, potential_predictor):
        """生成多Sheet的Excel报告"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        # 通用样式
        header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        good_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        warn_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        danger_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

        def write_sheet(ws, title, headers, data_rows):
            ws.title = title
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            for ri, row_data in enumerate(data_rows, 2):
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = thin_border
                    if isinstance(val, (int, float)):
                        cell.alignment = cell_align
                    else:
                        cell.alignment = left_align
            for ci in range(1, len(headers)+1):
                ws.column_dimensions[get_column_letter(ci)].width = max(12, len(str(headers[ci-1]))*2+4)
            ws.auto_filter.ref = ws.dimensions

        # Sheet 1: SKU推荐
        ws = wb.create_sheet('SKU投放推荐')
        top = sku_evaluator.get_top_skus(200)
        headers = ['排名','SKU','15天销量','15天销售额','投放评分','优先级','推荐广告','建议预算%']
        rows = []
        for i, (_, r) in enumerate(top.iterrows(), 1):
            rows.append([i, str(r['SKU']), r.get('15天销量',0), round(r.get('15天销售额',0),2),
                        r.get('投放评分',0), r.get('投放优先级',''), r.get('推荐广告类型',''),
                        round(r.get('建议预算占比',0)*100,1)])
        write_sheet(ws, 'SKU投放推荐', headers, rows)

        # Sheet 2: 关停建议
        ws = wb.create_sheet('关停广告词组')
        shutdown = keyword_optimizer.get_shutdown_list(limit=300)
        headers = ['序号','广告组','关键词','类型','点击','转化','ACoS','状态','原因']
        rows = []
        for i, (_, r) in enumerate(shutdown.iterrows(), 1):
            rows.append([i, str(r.get('Campaign',''))[:25], str(r.get('Keyword',''))[:25],
                        str(r.get('Match type',''))[:10], r.get('Clicks',0), r.get('Sales',0),
                        f"{r.get('ACOS',0):.1%}" if r.get('ACOS') else '-',
                        str(r.get('Status','')), str(r.get('优化建议',''))[:50]])
        write_sheet(ws, '关停广告词组', headers, rows)

        # Sheet 3: 扩量建议
        ws = wb.create_sheet('扩量广告词组')
        expand = keyword_optimizer.get_expand_list()
        headers = ['序号','广告组','关键词','类型','点击','转化','ACoS','ROAS','建议']
        rows = []
        for i, (_, r) in enumerate(expand.iterrows(), 1):
            rows.append([i, str(r.get('Campaign',''))[:25], str(r.get('Keyword',''))[:25],
                        str(r.get('Match type',''))[:10], r.get('Clicks',0), r.get('Sales',0),
                        f"{r.get('ACOS',0):.1%}" if r.get('ACOS') else '-',
                        f"{r.get('ROAS',0):.1f}" if r.get('ROAS') else '-',
                        str(r.get('优化建议',''))[:50]])
        write_sheet(ws, '扩量广告词组', headers, rows)

        # Sheet 3.5: 降价优化
        ws = wb.create_sheet('降价优化')
        optimize = keyword_optimizer.get_optimize_list()
        headers = ['序号','广告组','关键词','类型','点击','转化','ACoS','建议']
        rows = []
        for i, (_, r) in enumerate(optimize.iterrows(), 1):
            rows.append([i, str(r.get('Campaign',''))[:25], str(r.get('Keyword',''))[:25],
                        str(r.get('Match type',''))[:10], r.get('Clicks',0), r.get('Sales',0),
                        f"{r.get('ACOS',0):.1%}" if r.get('ACOS') else '-',
                        str(r.get('优化建议',''))[:50]])
        write_sheet(ws, '降价优化', headers, rows)

        # Sheet 4: 潜力SKU
        ws = wb.create_sheet('潜力SKU预判')
        high_pot = potential_predictor.get_high_potential_skus()
        headers = ['序号','SKU','15天销量','销量环比','潜力评分','潜力等级','分析']
        rows = []
        for i, (_, r) in enumerate(high_pot.head(100).iterrows(), 1):
            rows.append([i, str(r['SKU']), r.get('15天销量',0), str(r.get('销量环比','')),
                        r.get('潜力评分',0), r.get('潜力等级',''), str(r.get('潜力分析',''))[:60]])
        write_sheet(ws, '潜力SKU预判', headers, rows)

        wb.save(file_path)
        return file_path

    def save_html(self, file_path, sku_evaluator, keyword_optimizer, potential_predictor):
        """生成美观的HTML报告"""
        import json

        top_skus = sku_evaluator.get_top_skus(15)
        # 限制导出条数，避免Excel文件过大
        shutdown = keyword_optimizer.get_shutdown_list(limit=300)
        expand = keyword_optimizer.get_expand_list()
        optimize = keyword_optimizer.get_optimize_list()
        suggested = keyword_optimizer.get_suggested_keywords()
        high_pot = potential_predictor.get_high_potential_skus()

        from datetime import datetime
        now = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Build data arrays for JS
        sku_recs = []
        for i, (_, r) in enumerate(top_skus.iterrows(), 1):
            sku_recs.append({
                'rank': i, 'sku': str(r['SKU']), 'sales': int(r.get('15天销量',0)),
                'revenue': round(r.get('15天销售额',0),2), 'score': r.get('投放评分',0),
                'priority': r.get('投放优先级',''), 'ad_type': r.get('推荐广告类型',''),
                'budget': round(r.get('建议预算占比',0)*100,1)
            })

        shutdown_list = []
        for i, (_, r) in enumerate(shutdown.iterrows(), 1):
            shutdown_list.append({
                'rank': i, 'ad_group': str(r.get('Campaign',''))[:25],
                'keyword': str(r.get('Keyword',''))[:30],
                'match_type': str(r.get('Match type',''))[:10],
                'clicks': int(r.get('Clicks',0)),
                'acos': f"{r.get('ACOS',0):.1%}" if r.get('ACOS') else '-',
                'reason': str(r.get('优化建议',''))[:50]
            })

        expand_list = []
        for i, (_, r) in enumerate(expand.iterrows(), 1):
            expand_list.append({
                'rank': i, 'ad_group': str(r.get('Campaign',''))[:25],
                'keyword': str(r.get('Keyword',''))[:30],
                'match_type': str(r.get('Match type',''))[:10],
                'clicks': int(r.get('Clicks',0)),
                'acos': f"{r.get('ACOS',0):.1%}" if r.get('ACOS') else '-',
                'roas': f"{r.get('ROAS',0):.1f}" if r.get('ROAS') else '-',
                'advice': str(r.get('优化建议',''))[:50]
            })

        suggest_list = []
        for i, (_, r) in enumerate(suggested.head(15).iterrows(), 1):
            suggest_list.append({
                'rank': i, 'original': str(r.get('原始关键词',''))[:25],
                'suggested': str(r.get('建议关键词',''))[:35],
                'reason': str(r.get('建议原因',''))[:40]
            })

        pot_list = []
        for i, (_, r) in enumerate(high_pot.head(15).iterrows(), 1):
            pot_list.append({
                'rank': i, 'sku': str(r['SKU']), 'sales': int(r.get('15天销量',0)),
                'mom': str(r.get('销量环比','')), 'score': r.get('潜力评分',0),
                'level': r.get('潜力等级',''), 'analysis': str(r.get('潜力分析',''))[:60]
            })

        optimize_list = []
        for i, (_, r) in enumerate(optimize.iterrows(), 1):
            optimize_list.append({
                'rank': i, 'ad_group': str(r.get('Campaign',''))[:25],
                'keyword': str(r.get('Keyword',''))[:30],
                'match_type': str(r.get('Match type',''))[:10],
                'clicks': int(r.get('Clicks',0)),
                'sales': int(r.get('Sales',0)),
                'acos': f"{r.get('ACOS',0):.1%}" if r.get('ACOS') else '-',
                'advice': str(r.get('优化建议',''))[:50]
            })

        html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>eBay广告投放智能分析报告</title>
<style>
:root {{ --primary: #2B579A; --success: #4CAF50; --warning: #FF9800; --danger: #F44336; --bg: #F5F6FA; --card: #FFFFFF; --text: #333333; --muted: #888888; --border: #E0E0E0; }}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: 'Microsoft YaHei','PingFang SC',sans-serif; background:var(--bg); color:var(--text); line-height:1.6; }}
.container {{ max-width:1200px; margin:0 auto; padding:20px; }}
.header {{ background:linear-gradient(135deg, #1a3a6b 0%, #2B579A 100%); color:white; padding:40px 30px; border-radius:12px; margin-bottom:24px; }}
.header h1 {{ font-size:28px; margin-bottom:8px; }}
.header .meta {{ opacity:0.8; font-size:14px; }}
.card {{ background:var(--card); border-radius:12px; padding:24px; margin-bottom:20px; box-shadow:0 2px 8px rgba(0,0,0,0.06); }}
.card h2 {{ font-size:20px; color:var(--primary); margin-bottom:16px; padding-bottom:12px; border-bottom:2px solid var(--primary); }}
.card h3 {{ font-size:16px; color:#555; margin:16px 0 10px; }}
.metrics {{ display:flex; gap:16px; flex-wrap:wrap; margin-bottom:16px; }}
.metric {{ flex:1; min-width:160px; background:#F8F9FC; border-radius:8px; padding:16px; text-align:center; }}
.metric .value {{ font-size:24px; font-weight:700; color:var(--primary); }}
.metric .label {{ font-size:12px; color:var(--muted); margin-top:4px; }}
table {{ width:100%; border-collapse:collapse; margin:12px 0; font-size:13px; }}
th {{ background:var(--primary); color:white; padding:10px 12px; text-align:center; font-weight:600; white-space:nowrap; }}
td {{ padding:8px 12px; text-align:center; border-bottom:1px solid var(--border); }}
tr:hover {{ background:#F0F4FF; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:4px; font-size:11px; font-weight:600; }}
.badge-a {{ background:#E8F5E9; color:#2E7D32; }}
.badge-b {{ background:#E3F2FD; color:#1565C0; }}
.badge-c {{ background:#FFF3E0; color:#E65100; }}
.badge-d {{ background:#FFEBEE; color:#C62828; }}
.badge-shutdown {{ background:#FFEBEE; color:#C62828; }}
.badge-expand {{ background:#E8F5E9; color:#2E7D32; }}
.tips {{ background:#FFF9C4; border-left:4px solid #FBC02D; padding:16px 20px; border-radius:0 8px 8px 0; margin:16px 0; }}
.tips strong {{ color:#F57F17; }}
.footer {{ text-align:center; padding:24px; color:var(--muted); font-size:13px; }}
.grid {{ display:grid; grid-template-columns:1fr 1fr; gap:20px; }}
@media (max-width:768px) {{ .grid {{ grid-template-columns:1fr; }} }}
</style>
</head>
<body>
<div class="container">

<div class="header">
  <h1>📊 eBay广告投放智能分析报告</h1>
  <div class="meta">生成时间: {now} &nbsp;|&nbsp; 数据来源: 15天销售监测 + CPC广告数据 &nbsp;|&nbsp; 市场: 美国站</div>
</div>

<div class="card">
  <h2>一、eBay广告投放方法论</h2>
  <div class="tips">
    <strong>💡 核心原则：</strong>PLP(Advanced)按点击收费，适合新品测款、关键词探索，曝光不花钱留着也无妨；PLG(Standard)按成交收费，适合成熟商品收割，不出单不花钱。判断空烧的唯一标准是<strong>花了钱没产出</strong>，不看曝光量。
  </div>

  <h3>1. 广告类型对比</h3>
  <table>
    <tr><th>类型</th><th>付费方式</th><th>何时扣费</th><th>适用阶段</th><th>注意</th></tr>
    <tr><td><strong>PLG (Standard)</strong></td><td>按成交 CPS</td><td>出单才扣费</td><td>成熟/爆款期</td><td>佣金 12-15%，转化率 > 2% 再上</td></tr>
    <tr><td><strong>PLP (Advanced)</strong></td><td>按点击 CPC</td><td>点就扣费</td><td>新品/测款期</td><td>曝光不花钱，零点击不算浪费</td></tr>
  </table>

  <h3>2. 关键词四阶段决策标准</h3>
  <table>
    <tr><th>阶段</th><th>条件</th><th>判定</th><th>说明</th></tr>
    <tr><td>曝光期</td><td>曝光 &lt; 100</td><td><span class="badge badge-b">观察</span></td><td>刚上线或出价太低，等数据积累</td></tr>
    <tr><td>曝光→点击</td><td>曝光 ≥ 100，点击=0</td><td><span class="badge badge-b">观察</span></td><td>PLP 按点击收费，曝光不花钱，留着无妨</td></tr>
    <tr><td>点击→转化</td><td>点击 ≥ 15，零转化，花费 ≥ $5</td><td><span class="badge badge-shutdown">空烧</span></td><td>来了十几个人没一个买，钱白花了</td></tr>
    <tr><td>点击→转化</td><td>点击 &lt; 15，零转化</td><td><span class="badge badge-b">观察</span></td><td>样本不够，继续观察</td></tr>
    <tr><td>有转化</td><td>ACoS ≤ 15%</td><td><span class="badge badge-expand">追加预算</span></td><td>高效！提高出价 20-30%</td></tr>
    <tr><td>有转化</td><td>ACoS 15-30%</td><td>保持</td><td>健康区间，维持当前策略</td></tr>
    <tr><td>有转化</td><td>ACoS 30-60%</td><td><span class="badge badge-c">降价优化</span></td><td>微利，降低出价 15-25%</td></tr>
    <tr><td>有转化</td><td>ACoS &gt; 100%</td><td><span class="badge badge-shutdown">空烧</span></td><td>越投越亏，暂停或删除</td></tr>
  </table>
  <table>
    <tr><td>1.</td><td>从PLP开始：先用按点击付费测试关键词和商品</td></tr>
    <tr><td>2.</td><td>小预算起步：每个SKU每天预算$5-$10</td></tr>
    <tr><td>3.</td><td>密切监控：每天查看广告数据，及时调整</td></tr>
    <tr><td>4.</td><td>关键词精选：从广泛匹配开始，逐步筛选优质词</td></tr>
    <tr><td>5.</td><td>耐心测试：广告效果需时间积累，不要急于否定</td></tr>
  </table>
</div>

<div class="card">
  <h2>二、SKU投放价值评估 — Top {len(sku_recs)}</h2>
  <table>
    <tr><th>排名</th><th>SKU</th><th>15天销量</th><th>15天销售额</th><th>评分</th><th>优先级</th><th>推荐广告</th><th>预算%</th></tr>
'''
        for s in sku_recs:
            badge_cls = 'badge-a' if s['priority'] in ('A','高') else ('badge-b' if s['priority'] in ('B','中') else 'badge-c')
            html += f'    <tr><td>{s["rank"]}</td><td><strong>{s["sku"]}</strong></td><td>{s["sales"]}</td><td>${s["revenue"]:,.2f}</td><td>{s["score"]}</td><td><span class="badge {badge_cls}">{s["priority"]}</span></td><td>{s["ad_type"]}</td><td>{s["budget"]}%</td></tr>\n'

        html += f'''  </table>
</div>

<div class="card">
  <h2>三、广告词组优化建议</h2>
  <h3>3.1 建议空烧关停的广告词组 ({len(shutdown_list)}个)</h3>
'''
        if shutdown_list:
            html += '  <table>\n    <tr><th>序号</th><th>广告组</th><th>关键词</th><th>类型</th><th>点击</th><th>ACoS</th><th>原因</th></tr>\n'
            for s in shutdown_list:
                html += f'    <tr><td>{s["rank"]}</td><td>{s["ad_group"]}</td><td>{s["keyword"]}</td><td>{s["match_type"]}</td><td>{s["clicks"]}</td><td>{s["acos"]}</td><td>{s["reason"]}</td></tr>\n'
            html += '  </table>\n'
        else:
            html += '  <p>暂无需要关停的广告词组</p>\n'

        html += f'  <h3>3.2 建议追加预算的广告词组 ({len(expand_list)}个)</h3>\n'
        if expand_list:
            html += '  <table>\n    <tr><th>序号</th><th>广告组</th><th>关键词</th><th>类型</th><th>点击</th><th>ACoS</th><th>ROAS</th><th>建议</th></tr>\n'
            for s in expand_list:
                html += f'    <tr><td>{s["rank"]}</td><td>{s["ad_group"]}</td><td>{s["keyword"]}</td><td>{s["match_type"]}</td><td>{s["clicks"]}</td><td>{s["acos"]}</td><td>{s["roas"]}</td><td>{s["advice"]}</td></tr>\n'
            html += '  </table>\n'
        else:
            html += '  <p>暂无需要扩量的广告词组</p>\n'

        html += f'  <h3>3.3 建议降价的广告词组 ({len(optimize_list)}个)</h3>\n'
        if optimize_list:
            html += '  <table>\n    <tr><th>序号</th><th>广告组</th><th>关键词</th><th>类型</th><th>点击</th><th>转化</th><th>ACoS</th><th>建议</th></tr>\n'
            for s in optimize_list:
                html += f'    <tr><td>{s["rank"]}</td><td>{s["ad_group"]}</td><td>{s["keyword"]}</td><td>{s["match_type"]}</td><td>{s["clicks"]}</td><td>{s["sales"]}</td><td>{s["acos"]}</td><td>{s["advice"]}</td></tr>\n'
            html += '  </table>\n'
        else:
            html += '  <p>暂无需要降价的广告词组</p>\n'

        html += f'  <h3>3.4 建议新增的广告词组 ({len(suggest_list)}个)</h3>\n'
        if suggest_list:
            html += '  <table>\n    <tr><th>序号</th><th>原始关键词</th><th>建议关键词</th><th>原因</th></tr>\n'
            for s in suggest_list:
                html += f'    <tr><td>{s["rank"]}</td><td>{s["original"]}</td><td><strong>{s["suggested"]}</strong></td><td>{s["reason"]}</td></tr>\n'
            html += '  </table>\n'
        else:
            html += '  <p>暂无建议新增的关键词</p>\n'

        html += f'''</div>

<div class="card">
  <h2>四、SKU投放潜力预判</h2>
  <h3>4.1 高潜力SKU Top {len(pot_list)}</h3>
  <table>
    <tr><th>排名</th><th>SKU</th><th>15天销量</th><th>环比</th><th>潜力评分</th><th>等级</th><th>分析</th></tr>
'''
        for s in pot_list:
            badge_cls = 'badge-a' if s['level'] in ('高','A') else ('badge-b' if s['level'] in ('中','B') else 'badge-c')
            html += f'    <tr><td>{s["rank"]}</td><td><strong>{s["sku"]}</strong></td><td>{s["sales"]}</td><td>{s["mom"]}</td><td>{s["score"]}</td><td><span class="badge {badge_cls}">{s["level"]}</span></td><td style="text-align:left">{s["analysis"]}</td></tr>\n'

        html += f'''</div>

<div class="card">
  <h2>五、总结与建议</h2>
  <div class="grid">
    <div>
      <h3>▶ 立即执行</h3>
      <p>关停零转化的广告词组，对高优先级SKU增加广告预算</p>
    </div>
    <div>
      <h3>▶ 本周执行</h3>
      <p>测试新推荐的广告词组，关注新品机会的投放效果</p>
    </div>
    <div>
      <h3>▶ 持续优化</h3>
      <p>每日查看广告数据，根据ACoS调整出价策略</p>
    </div>
    <div>
      <h3>⚠️ 常见误区</h3>
      <p>不给所有商品都投广告；不忽略ACoS超50%；数据至少观察7天</p>
    </div>
  </div>
</div>

<div class="footer">
  eBay广告投放智能分析系统 &copy; 2026 | 报告自动生成
</div>

</div>
</body>
</html>'''

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return file_path
